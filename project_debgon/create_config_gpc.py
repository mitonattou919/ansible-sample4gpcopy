#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789

# coding: utf-8

import openpyxl
import sys
import os

# Define Worksheet name.
ws_name_env     = 'env'
ws_name_files   = 'files'

# Define file name.
inventory_file = 'hosts'
var_dir_host   = 'host_vars'
var_dir_group  = 'group_vars'

max_target = 16                # Max number of target host.

# [ START write_file]
# Write strings to file.
#
# Args:
#   file_name: specify target file name.
#   text: text strings to write.
# Returns:
#   none
#
#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
def write_file(file_name, text):
    with open(file_name, mode='a') as f:
        f.write(text)

# [ END write_file]


# Get target Excel file name.
args = sys.argv
my_wb = openpyxl.load_workbook(args[1], data_only=True)

ws_env     = my_wb[ws_name_env]
ws_files   = my_wb[ws_name_files]

# Environment type.
## Get environment type.
env_type = ws_env.cell(2, 2).value

## Confirm exectuion environment is correct.
question_msg = ('You execute this script in ' + env_type 
                + ' environment.\nAre you sure? [y/N]: ')
user_answer = input(question_msg)

if user_answer != 'y':
    exit()



#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Get environment informations.
#
env_values = []

for col in ws_env.iter_cols(min_row=5, min_col=2, max_row=9):
    env_values.append([cell.value for cell in col])
    #if env_values[0][0] is None: break


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Inventory file.
#
## File initialization.
if os.path.exists(inventory_file + '_' + env_type):
    os.remove(inventory_file + '_' + env_type)

## Create inventory file
for env_value in env_values:
    if env_value[0] != None:
        text = ('[' + str(env_value[1]) + ']\n' 
                + str(env_value[1]) 
                + ' ansible_host=' + str(env_value[2])
                + ' ansible_user=' + str(env_value[3])
                + ' ansible_password=' + str(env_value[4]) + '\n\n')
        write_file(inventory_file + '_' + env_type, text)


#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Variable files.
#
## Directory initialization.
if not os.path.exists(var_dir_host):
    os.makedirs('./' + var_dir_host)



#         1         2         3         4         5         6         7
#1234567890123456789012345678901234567890123456789012345678901234567890123456789
#
# Files
#
## Get directory informations.
#
values = []

for row in ws_files.iter_rows(min_row=1, min_col=1):
    values.append([cell.value for cell in row])

for value in values:
    if value[0] == 'path':
        for i in range(max_target):
            if i >= 4 and value[i] is not None:
                if value[i] == env_values[i - 4][0]:
                    if os.path.exists(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml'):
                        os.remove(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml')

                    text = ('\n# General purpose copy\n'
                            + 'general_purpose_copy_info:\n')
                    write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
                else:
                    print('Error: Target host is not match!')
                    exit()
    else:
        for i in range(max_target):
            if i >= 4 and value[i] is not None and value[i] == 'yes':
                text = (' - { target_file: \'' + str(value[0])[1:] + '\', '
                        + 'owner: \'' + str(value[1]) + '\', ' 
                        + 'group: \'' + str(value[2]) + '\', '
                        + 'mode: \'' + str(value[3]) + '\' }\n')
                write_file(var_dir_host + '/' + str(env_values[i - 4][0]) + '.yml', text)
